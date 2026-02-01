#!/usr/bin/env python3
"""
Optuna-driven hyperparameter search with the same run/update behavior as main_expe_overall.py.
Runs PPO-EDA (decay mode) over all QUBO/NK instances and updates results/config/overall_summary.xlsx.
Skips configs already evaluated.
"""

from __future__ import annotations

import argparse
import os
from pathlib import Path

try:
    import optuna
except Exception as exc:
    raise SystemExit("Optuna is required. Install it in your venv: pip install optuna") from exc

from openpyxl import load_workbook

from main_expe_overall import (
    DEFAULTS,
    _build_config_name,
    _collect_config_stats,
    _discover_nk_instances,
    _discover_qubo_instances,
    _is_cuda_oom,
    _load_instances,
    _update_excel_summary,
    _run_once,
)


# Fixed params
KERNEL = "jsd"
ADVANTAGE = "peragentrankweighted"
GAMMA = 0.0005
BANDWITH_KERNEL = None

# Search space (edit here)
M_VALUES = [3, 4, 5, 6]
LAMBDA_VALUES = [20, 24, 30]
EPSILON_VALUES = [0.005, 0.007, 0.01]
DECAY_START_VALUES = [0.05, 0.1, 0.2]
DECAY_MIN_VALUES = [0.05, 0.1, 0.2]


def _load_seen_configs(xlsx_path: str) -> set[str]:
    if not os.path.isfile(xlsx_path):
        return set()
    wb = load_workbook(xlsx_path)
    if "summary" not in wb.sheetnames:
        return set()
    ws = wb["summary"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return set()
    header = [str(h) if h is not None else "" for h in rows[0]]
    if "config_name" not in header:
        return set()
    idx = header.index("config_name")
    out = set()
    for r in rows[1:]:
        if not r or idx >= len(r):
            continue
        cfg = r[idx]
        if cfg:
            out.add(str(cfg))
    return out


def _config_params_from_trial(trial: optuna.Trial) -> dict:
    return dict(
        kernel=KERNEL,
        advantage=ADVANTAGE,
        M=trial.suggest_categorical("M", M_VALUES),
        lambda_=trial.suggest_categorical("lambda_", LAMBDA_VALUES),
        epsilon_svgd=trial.suggest_categorical("epsilon_svgd", EPSILON_VALUES),
        gamma=GAMMA,
        decay_start_ratio=trial.suggest_categorical("decay_start_ratio", DECAY_START_VALUES),
        decay_min_factor=trial.suggest_categorical("decay_min_factor", DECAY_MIN_VALUES),
        bandwith_kernel=BANDWITH_KERNEL,
    )


def _score_from_stats(stats: dict) -> float:
    # Lexicographic preference: top1_count desc, median_rank asc, mean_hamming_norm desc.
    top1 = stats.get("top1_count") or 0
    median_rank = stats.get("median_rank")
    mean_hamming = stats.get("mean_hamming_norm") or 0.0
    median_rank_val = float(median_rank) if median_rank is not None else 1e9
    return top1 * 1e6 - median_rank_val * 1e3 + mean_hamming


def main():
    parser = argparse.ArgumentParser(description="Optuna search for PPO-EDA hyperparams.")
    parser.add_argument("--trials", type=int, default=10, help="Number of Optuna trials to run.")
    parser.add_argument("--study", type=str, default="optuna_overall", help="Optuna study name.")
    parser.add_argument("--outdir", type=str, default=None, help="Root output dir (default: results/config).")
    args = parser.parse_args()

    repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    out_root = args.outdir or os.path.join(repo_root, "results", "config")
    Path(out_root).mkdir(parents=True, exist_ok=True)

    out_xlsx = os.path.join(out_root, "overall_summary.xlsx")
    seen_configs = _load_seen_configs(out_xlsx)

    instances_root = Path(repo_root) / "source_code" / "instances"
    qubo_instances = _discover_qubo_instances(instances_root / "QUBO", DEFAULTS["nb_instances_test"])
    nk_instances = _discover_nk_instances(instances_root / "nk", DEFAULTS["nb_instances_test"])
    instances = qubo_instances + nk_instances
    if not instances:
        raise SystemExit("Aucune instance QUBO/NK compatible avec nb_instances_test.")

    storage = f"sqlite:///{os.path.join(out_root, 'optuna_study.db')}"
    study = optuna.create_study(
        study_name=args.study,
        direction="maximize",
        sampler=optuna.samplers.TPESampler(seed=DEFAULTS["seed"]),
        storage=storage,
        load_if_exists=True,
    )

    for idx in range(1, args.trials + 1):
        trial = study.ask()
        params = _config_params_from_trial(trial)
        config_name = _build_config_name(None, params)
        print(f"[TRIAL {idx}/{args.trials}] {config_name}")
        if config_name in seen_configs:
            print("  -> skip config (already evaluated)")
            study.tell(trial, float("-inf"))
            continue

        config_dir = os.path.join(out_root, config_name)
        pending_instances = []
        skipped_instances = []
        for inst in instances:
            inst_name = f"{inst['name']}_dim{inst['dim']}_t{inst['type_instance']}"
            inst_dir = os.path.join(config_dir, inst_name)
            summary_path = os.path.join(inst_dir, "best_summary.txt")
            if os.path.isfile(summary_path):
                skipped_instances.append(inst_name)
            else:
                pending_instances.append(inst)

        for inst_name in skipped_instances:
            print(f"  -> skip {inst_name} (already done)")

        for inst in pending_instances:
            inst_name = f"{inst['name']}_dim{inst['dim']}_t{inst['type_instance']}"
            inst_dir = os.path.join(config_dir, inst_name)
            problem_ctx = _load_instances(inst, DEFAULTS["device"])
            print(f"  -> run {inst_name}")

            nb_restarts = DEFAULTS["nb_restarts"]
            success = False
            while nb_restarts > 0 and not success:
                try:
                    avg_score, history, meta = _run_once(
                        problem_ctx,
                        params["kernel"],
                        params["advantage"],
                        params["M"],
                        params["lambda_"],
                        params["epsilon_svgd"],
                        params["gamma"],
                        params["decay_start_ratio"],
                        params["decay_min_factor"],
                        params.get("bandwith_kernel"),
                        device=DEFAULTS["device"],
                        nb_restarts=nb_restarts,
                    )
                    success = True
                except (RuntimeError, Exception) as exc:
                    if not _is_cuda_oom(exc):
                        raise
                    nb_restarts -= 1
                    if nb_restarts <= 0:
                        break

            if not success:
                continue

            from main_expe_overall import _save_history_csv, _rank_vs_global_ranking_excluding_ppo

            ranking = _rank_vs_global_ranking_excluding_ppo(
                repo_root, inst["name"], inst["dim"], inst["type_instance"], avg_score
            )
            if ranking and ranking[2] == 1:
                print("     -> TOP 1")
            _save_history_csv(
                inst_dir,
                inst["name"],
                params["kernel"],
                {"history": history, "meta": meta},
                ranking=ranking,
                config_name=config_name,
            )
            _update_excel_summary(out_xlsx, config_name, params, config_dir, repo_root)

        stats = _collect_config_stats(config_dir, config_name, params, repo_root)
        score = _score_from_stats(stats)
        study.tell(trial, score)
        seen_configs.add(config_name)

    print(f"[DONE] Optuna trials complete. Summary: {out_xlsx}")


if __name__ == "__main__":
    main()
