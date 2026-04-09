"""
Run PPO-EDA over all QUBO/NK instances using the same grid logic as main_expe_overall,
but with SVGD repulsion disabled (no_repulsion=True).
Results are stored under:
  results/config/<ConfigName>/<InstanceName>/no_repulsion/
"""

from __future__ import annotations

import argparse
import os
import sys
import time
from pathlib import Path

SOURCE_CODE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if SOURCE_CODE_DIR not in sys.path:
    sys.path.insert(0, SOURCE_CODE_DIR)

import numpy as np
import torch
from openpyxl import Workbook, load_workbook

from eda_strategies.FactoryStrategyEA import FactoryStrategyEA
from environment.nk import get_Score_trajectoriesNK_cuda, getTensorInstances_NK
from environment.qubo import get_Score_trajectoriesQUBO_cuda
from expe.main_expe_overall import (
    DEFAULTS,
    INSTANCE_DIR_RE,
    _discover_nk_instances,
    _discover_qubo_instances,
    _format_float,
    _is_cuda_oom,
    _load_instances,
    _parse_summary_config,
    _rank_vs_global_ranking_excluding_ppo,
    _round_float,
    _save_history_csv,
    _set_seeds,
)

DEFAULT_CONFIG_NAME = "krbf__advglobalrankweighted__M7__L13__eps0p08__g0p015__ds0p03__dm0p01"

SUMMARY_COLUMNS = [
    "config_name",
    "kernel",
    "advantage",
    "M",
    "lambda_",
    "epsilon_svgd",
    "gamma",
    "decay_start_ratio",
    "decay_min_factor",
    "mean_rank",
    "median_rank",
    "std_percent",
    "top1_count",
    "top3_count",
    "top5_count",
    "top10_count",
    "top_1_nk",
    "top_1_qubo",
    "win_rate_mean",
    "mean_hamming_norm",
    "mean_l1_norm",
    "n_instances",
    "n_ranked",
]


def _load_summary_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}, []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    if not any(header):
        return {}, header
    if "config_name" not in header:
        return {}, header
    idx = {name: i for i, name in enumerate(header) if name}
    rows_dict = {}
    for row in rows[1:]:
        if not row:
            continue
        name = row[idx["config_name"]] if "config_name" in idx and idx["config_name"] < len(row) else None
        if not name:
            continue
        entry = {}
        for col in SUMMARY_COLUMNS:
            if col in idx and idx[col] < len(row):
                entry[col] = row[idx[col]]
        rows_dict[str(name)] = entry
    return rows_dict, header


def _write_summary_sheet(ws, rows_dict):
    ws.delete_rows(1, ws.max_row)
    ws.append(SUMMARY_COLUMNS)
    for name in sorted(rows_dict.keys()):
        entry = dict(rows_dict[name] or {})
        entry.setdefault("config_name", name)
        ws.append([entry.get(col, "") for col in SUMMARY_COLUMNS])

def _parse_config_name(config_name: str) -> dict:
    parts = config_name.split("__")
    out = {
        "kernel": None,
        "advantage": None,
        "M": None,
        "lambda_": None,
        "epsilon_svgd": None,
        "gamma": None,
        "decay_start_ratio": None,
        "decay_min_factor": None,
        "bandwith_kernel": None,
    }

    def parse_float(token: str):
        token = token.replace("p", ".").replace("m", "-")
        try:
            return float(token)
        except Exception:
            return None

    for p in parts:
        if p.startswith("k"):
            out["kernel"] = p[1:]
        elif p.startswith("adv"):
            out["advantage"] = p[3:]
        elif p.startswith("M"):
            try:
                out["M"] = int(p[1:])
            except Exception:
                pass
        elif p.startswith("L"):
            try:
                out["lambda_"] = int(p[1:])
            except Exception:
                pass
        elif p.startswith("eps"):
            out["epsilon_svgd"] = parse_float(p[3:])
        elif p.startswith("g"):
            out["gamma"] = parse_float(p[1:])
        elif p.startswith("ds"):
            out["decay_start_ratio"] = parse_float(p[2:])
        elif p.startswith("dm"):
            out["decay_min_factor"] = parse_float(p[2:])
        elif p.startswith("bw"):
            out["bandwith_kernel"] = parse_float(p[2:])
    return out


def _run_once_no_repulsion(
    problem_ctx,
    kernel_name,
    advantage,
    M,
    lambda_,
    epsilon_svgd,
    gamma,
    decay_start_ratio,
    decay_min_factor,
    bandwith_kernel,
    device=None,
    nb_restarts=None,
):
    device = device or DEFAULTS["device"]
    nb_restarts = DEFAULTS["nb_restarts"] if nb_restarts is None else int(nb_restarts)

    kernel_config = {"name": kernel_name, "epsilon_svgd": epsilon_svgd, "gamma": gamma}
    if kernel_name in ("rbf", "pk") and bandwith_kernel is not None:
        kernel_config["bandwith_kernel"] = bandwith_kernel

    factory = FactoryStrategyEA()
    strategy = factory.createStrategyEA(
        "PPO-EDA",
        problem_ctx["dim"],
        lambda_,
        device,
        problem_ctx["dim_variables"],
        M,
        learning_rate=epsilon_svgd,
        epsilon_svgd=epsilon_svgd,
        enable_visualization=DEFAULTS["visualization"],
        svgd_gamma=gamma,
        decay_start_ratio=decay_start_ratio,
        decay_min_factor=decay_min_factor,
        decay_enabled=True,
        advantage_cfg=advantage,
        kernel_config=kernel_config,
        no_interact=False,
        no_repulsion=True,
    ).to(device)

    if problem_ctx["type_problem"] == "QUBO":
        list_scores, history = get_Score_trajectoriesQUBO_cuda(
            strategy,
            problem_ctx["dim"],
            DEFAULTS["nb_instances_test"],
            nb_restarts,
            DEFAULTS["budget"],
            lambda_,
            problem_ctx["tensor_Q_test"],
            device,
            False,
            enable_visualization=False,
            return_history=True,
        )
    else:
        total_lambda = strategy.lambda_
        tensor_matrix_locus, tensor_matrix_contrib, _ = getTensorInstances_NK(
            problem_ctx["nk_base_path"],
            DEFAULTS["nb_instances_test"],
            nb_restarts,
            total_lambda,
            problem_ctx["dim"],
            problem_ctx["D"],
            problem_ctx["type_instance"],
            device,
        )
        list_scores, history = get_Score_trajectoriesNK_cuda(
            strategy,
            problem_ctx["dim"],
            problem_ctx["type_instance"],
            problem_ctx["D"],
            DEFAULTS["nb_instances_test"],
            nb_restarts,
            DEFAULTS["budget"],
            total_lambda,
            problem_ctx["vectorIndex_th"],
            tensor_matrix_locus,
            tensor_matrix_contrib,
            device,
            False,
            enable_visualization=False,
            return_history=True,
        )

    scores_array = (
        np.asarray(list_scores)
        if isinstance(list_scores, (list, tuple, np.ndarray))
        else (list_scores.detach().cpu().numpy() if torch.is_tensor(list_scores) else np.asarray(list_scores))
    )
    avg_score = float(np.mean(scores_array))
    median_score = float(np.percentile(scores_array, 50))
    std_score = float(np.std(scores_array))
    p2 = float(np.percentile(scores_array, 2))
    p5 = float(np.percentile(scores_array, 5))
    p10 = float(np.percentile(scores_array, 10))
    p25 = float(np.percentile(scores_array, 25))
    p50 = float(np.percentile(scores_array, 50))
    p75 = float(np.percentile(scores_array, 75))
    p90 = float(np.percentile(scores_array, 90))
    p95 = float(np.percentile(scores_array, 95))
    p98 = float(np.percentile(scores_array, 98))

    run_meta = dict(
        problem=problem_ctx["type_problem"],
        dim=problem_ctx["dim"],
        type_instance=problem_ctx["type_instance"],
        kernel=kernel_name,
        advantage=advantage,
        M=M,
        lambda_=lambda_,
        epsilon_svgd=epsilon_svgd,
        gamma=gamma,
        decay_start_ratio=decay_start_ratio,
        decay_min_factor=decay_min_factor,
        bandwith_kernel=bandwith_kernel,
        no_interact=False,
        no_repulsion=True,
        avg_score=avg_score,
        median_score=median_score,
        std_score=std_score,
        p2=p2,
        p5=p5,
        p10=p10,
        p25=p25,
        p50=p50,
        p75=p75,
        p90=p90,
        p95=p95,
        p98=p98,
    )
    return avg_score, history, run_meta


def _save_history_csv_no_repulsion(out_dir, problem_name, kernel_name, entry, ranking=None, config_name=None):
    _save_history_csv(
        out_dir,
        problem_name,
        kernel_name,
        entry,
        ranking=ranking,
        config_name=config_name,
    )
    summary_path = Path(out_dir) / "best_summary.txt"
    if not summary_path.is_file():
        return
    try:
        with open(summary_path, "r") as f:
            content = f.read()
        if "no_repulsion:" in content:
            return
        with open(summary_path, "a") as f:
            f.write("no_repulsion: True\n")
    except OSError:
        return


def _read_last_metric_score(metrics_path: Path):
    """Return (score, metric_name) from last line. Prefer mean, then median, then best_fitness."""
    try:
        lines = [line.strip() for line in metrics_path.read_text().splitlines() if line.strip()]
        if len(lines) < 2:
            return None, None
        header = [h.strip() for h in lines[0].split(",")]
        last = [v.strip() for v in lines[-1].split(",")]

        def pick(col: str):
            if col in header:
                idx = header.index(col)
                if idx < len(last) and last[idx] != "":
                    return last[idx]
            return None

        for col in ("mean", "median", "best_fitness"):
            val = pick(col)
            if val is not None:
                return float(val), col
        return None, None
    except Exception:
        return None, None


def _collect_config_stats_no_repulsion(config_dir: str, config_name: str, params: dict, repo_root: str):
    rows = []
    config_path = Path(config_dir)
    if not config_path.is_dir():
        return dict(
            config_name=config_name,
            kernel=params["kernel"],
            advantage=params["advantage"],
            M=params["M"],
            lambda_=params["lambda_"],
            epsilon_svgd=_round_float(params["epsilon_svgd"]),
            gamma=_round_float(params["gamma"]),
            decay_start_ratio=_round_float(params["decay_start_ratio"]),
            decay_min_factor=_round_float(params["decay_min_factor"]),
            mean_rank=None,
            median_rank=None,
            std_percent=None,
            top1_count=0,
            top3_count=0,
            top5_count=0,
            top10_count=0,
            top_1_nk=0,
            top_1_qubo=0,
            win_rate_mean=None,
            mean_hamming_norm=None,
            mean_l1_norm=None,
            n_instances=0,
            n_ranked=0,
        )

    for child in sorted(config_path.iterdir()):
        if not child.is_dir():
            continue
        match = INSTANCE_DIR_RE.match(child.name)
        if not match:
            continue
        problem = match.group("problem")
        dim = int(match.group("dim"))
        t = int(match.group("t"))

        run_dir = child / "no_repulsion"
        metrics_path = run_dir / "best_metrics.csv"
        if not metrics_path.is_file():
            continue
        avg_score, _metric_name = _read_last_metric_score(metrics_path)
        if avg_score is None:
            continue

        best_algo, best_score, my_rank, n_rank, my_pct, _, _ = _rank_vs_global_ranking_excluding_ppo(
            repo_root, problem, dim, t, avg_score
        )
        win_rate = None
        if my_rank is not None and n_rank:
            win_rate = (n_rank - my_rank) / n_rank

        hamming_norm = None
        l1_norm = None
        try:
            with open(metrics_path, "r") as f:
                lines = [line.strip() for line in f.readlines() if line.strip()]
            if len(lines) >= 2:
                header = lines[0].split(",")
                last = lines[-1].split(",")
                idx_ham = header.index("avg_hamming") if "avg_hamming" in header else None
                idx_l1 = header.index("avg_l1") if "avg_l1" in header else None
                if idx_ham is not None and idx_ham < len(last):
                    hamming_val = float(last[idx_ham])
                    hamming_norm = hamming_val / dim if hamming_val > 1 else hamming_val
                if idx_l1 is not None and idx_l1 < len(last):
                    l1_val = float(last[idx_l1])
                    l1_norm = l1_val / dim if l1_val > 1 else l1_val
        except Exception:
            hamming_norm = None
            l1_norm = None

        rows.append(
            dict(
                problem=problem,
                rank=my_rank,
                percent=my_pct,
                top1_count=1 if my_rank == 1 else 0,
                top3_count=1 if my_rank is not None and my_rank <= 3 else 0,
                top5_count=1 if my_rank is not None and my_rank <= 5 else 0,
                top10_count=1 if my_rank is not None and my_rank <= 10 else 0,
                ranking_best_algo=best_algo,
                ranking_best_score=best_score,
                n_rank=n_rank,
                win_rate=win_rate,
                hamming_norm=hamming_norm,
                l1_norm=l1_norm,
            )
        )

    n_instances = len(rows)
    ranks = [r["rank"] for r in rows if r["rank"] is not None]
    percents = [r["percent"] for r in rows if r["percent"] is not None]
    win_rates = [r["win_rate"] for r in rows if r.get("win_rate") is not None]
    hamming_vals = [r["hamming_norm"] for r in rows if r.get("hamming_norm") is not None]
    l1_vals = [r["l1_norm"] for r in rows if r.get("l1_norm") is not None]
    n_ranked = len(ranks)

    mean_rank = float(np.mean(ranks)) if ranks else None
    median_rank = float(np.median(ranks)) if ranks else None
    std_percent = float(np.std(percents)) if len(percents) > 1 else 0.0 if percents else None
    top1_count = sum(1 for r in rows if r.get("top1_count"))
    top3_count = sum(1 for r in rows if r.get("top3_count"))
    top5_count = sum(1 for r in rows if r.get("top5_count"))
    top10_count = sum(1 for r in rows if r.get("top10_count"))
    top_1_nk = sum(1 for r in rows if r.get("top1_count") and r.get("problem") == "NK")
    top_1_qubo = sum(1 for r in rows if r.get("top1_count") and r.get("problem") == "QUBO")
    win_rate_mean = float(np.mean(win_rates)) if win_rates else None
    mean_hamming_norm = float(np.mean(hamming_vals)) if hamming_vals else None
    mean_l1_norm = float(np.mean(l1_vals)) if l1_vals else None

    return dict(
        config_name=config_name,
        kernel=params["kernel"],
        advantage=params["advantage"],
        M=params["M"],
        lambda_=params["lambda_"],
        epsilon_svgd=_round_float(params["epsilon_svgd"]),
        gamma=_round_float(params["gamma"]),
        decay_start_ratio=_round_float(params["decay_start_ratio"]),
        decay_min_factor=_round_float(params["decay_min_factor"]),
        mean_rank=mean_rank,
        median_rank=median_rank,
        std_percent=std_percent,
        top1_count=top1_count,
        top3_count=top3_count,
        top5_count=top5_count,
        top10_count=top10_count,
        top_1_nk=top_1_nk,
        top_1_qubo=top_1_qubo,
        win_rate_mean=win_rate_mean,
        mean_hamming_norm=mean_hamming_norm,
        mean_l1_norm=mean_l1_norm,
        n_instances=n_instances,
        n_ranked=n_ranked,
    )


def _update_excel_summary_no_repulsion(out_xlsx: str, config_name: str, params: dict, config_dir: str, repo_root: str):
    stats = _collect_config_stats_no_repulsion(config_dir, config_name, params, repo_root)
    if stats["n_instances"] == 0:
        return

    if os.path.isfile(out_xlsx):
        wb = load_workbook(out_xlsx)
        if "summary" in wb.sheetnames:
            ws = wb["summary"]
            rows_dict, _ = _load_summary_rows(ws)
        else:
            ws = wb.create_sheet("summary")
            rows_dict = {}
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "summary"
        rows_dict = {}

    rows_dict[config_name] = stats
    _write_summary_sheet(ws, rows_dict)
    wb.save(out_xlsx)


def main():
    parser = argparse.ArgumentParser(description="Overall PPO-EDA grid with no_repulsion=True (QUBO + NK).")
    parser.add_argument("--outdir", type=str, default=None, help="Root output dir (default: results/config).")
    args = parser.parse_args()
    config_name = (
        input(
            f"Config name to test (ex: {DEFAULT_CONFIG_NAME}) [default: {DEFAULT_CONFIG_NAME}]: "
        ).strip()
        or DEFAULT_CONFIG_NAME
    )
    params = _parse_config_name(config_name)
    missing = [k for k, v in params.items() if v is None and k != "bandwith_kernel"]
    if missing:
        raise SystemExit(f"Invalid config_name, missing: {', '.join(missing)}")

    repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    out_root = args.outdir or os.path.join(repo_root, "results", "config")
    Path(out_root).mkdir(parents=True, exist_ok=True)
    instances_root = Path(repo_root) / "source_code" / "instances"
    qubo_instances = _discover_qubo_instances(instances_root / "QUBO", DEFAULTS["nb_instances_test"])
    nk_instances = _discover_nk_instances(instances_root / "nk", DEFAULTS["nb_instances_test"])
    instances = qubo_instances + nk_instances
    if not instances:
        raise SystemExit("Aucune instance QUBO/NK compatible avec nb_instances_test.")

    _set_seeds(DEFAULTS["seed"])

    out_xlsx = os.path.join(out_root, "overall_summary_no_repulsion.xlsx")
    start_all = time.time()
    config_dir = os.path.join(out_root, config_name)
    print(f"[CONFIG no_repulsion] {config_name}")

    _update_excel_summary_no_repulsion(out_xlsx, config_name, params, config_dir, repo_root)

    pending_instances = []
    skipped_instances = []
    for inst in instances:
        inst_name = f"{inst['name']}_dim{inst['dim']}_t{inst['type_instance']}"
        run_dir = os.path.join(config_dir, inst_name, "no_repulsion")
        summary_path = os.path.join(run_dir, "best_summary.txt")
        if os.path.isfile(summary_path):
            cfg = _parse_summary_config(Path(summary_path))
            avg_score = cfg.get("avg_score") if cfg else None
            if avg_score is not None:
                skipped_instances.append(inst_name)
                continue
        pending_instances.append(inst)

    for inst_name in skipped_instances:
        print(f"  -> skip {inst_name} (already done in no_repulsion)")

    if not pending_instances:
        print("  -> already complete for no_repulsion, skipping.")

    for inst in pending_instances:
        inst_name = f"{inst['name']}_dim{inst['dim']}_t{inst['type_instance']}"
        run_dir = os.path.join(config_dir, inst_name, "no_repulsion")
        problem_ctx = _load_instances(inst, DEFAULTS["device"])
        print(f"  -> run {inst_name} [no_repulsion]")
        t0 = time.time()
        nb_restarts = DEFAULTS["nb_restarts"]
        if inst["name"] == "NK" and inst["dim"] >= 256 and inst["type_instance"] >= 8:
            nb_restarts = min(nb_restarts, 3)
            print(f"     [PRE] NK big instance, start nb_restarts={nb_restarts}")
        success = False
        while nb_restarts > 0 and not success:
            try:
                avg_score, history, meta = _run_once_no_repulsion(
                    problem_ctx,
                    params["kernel"],
                    params["advantage"],
                    params["M"],
                    params["lambda_"],
                    params["epsilon_svgd"],
                    params["gamma"],
                    params["decay_start_ratio"],
                    params["decay_min_factor"],
                    params.get("bandwith_kernel"),
                    device=DEFAULTS["device"],
                    nb_restarts=nb_restarts,
                )
                success = True
            except (torch.OutOfMemoryError, RuntimeError) as exc:
                if not _is_cuda_oom(exc):
                    raise
                nb_restarts -= 1
                if torch.cuda.is_available():
                    torch.cuda.empty_cache()
                if nb_restarts > 0:
                    print(f"     [OOM] retry with nb_restarts={nb_restarts}.")
                else:
                    print("     [OOM] nb_restarts=0, skip instance.")

        if not success:
            continue
        dt = time.time() - t0
        print(f"     avg_score={_format_float(avg_score, 6)} | runtime={dt:.2f}s")
        ranking = _rank_vs_global_ranking_excluding_ppo(
            repo_root, inst["name"], inst["dim"], inst["type_instance"], avg_score
        )
        _save_history_csv_no_repulsion(
            run_dir,
            inst["name"],
            params["kernel"],
            {"history": history, "meta": meta},
            ranking=ranking,
            config_name=config_name,
        )
        if ranking and ranking[2] == 1:
            print("     -> TOP 1")
        _update_excel_summary_no_repulsion(out_xlsx, config_name, params, config_dir, repo_root)

    gap_lines = []
    wins_normal = 0
    wins_no_repulsion = 0
    gap_values = []
    ranks_normal = []
    ranks_no_repulsion = []

    for inst in instances:
        inst_name = f"{inst['name']}_dim{inst['dim']}_t{inst['type_instance']}"
        normal_metrics = Path(config_dir) / inst_name / "best_metrics.csv"
        no_repulsion_metrics = Path(config_dir) / inst_name / "no_repulsion" / "best_metrics.csv"
        if not normal_metrics.exists() or not no_repulsion_metrics.exists():
            continue
        normal_score, _ = _read_last_metric_score(normal_metrics)
        nr_score, _ = _read_last_metric_score(no_repulsion_metrics)
        if normal_score is None or nr_score is None:
            continue

        def _normalize_score_sign_local(problem_name: str, values):
            if problem_name.upper() in ("QUBO", "UBQP"):
                return [-val for val in values]
            return values

        norm_normal = _normalize_score_sign_local(inst["name"], [normal_score])[0]
        norm_nr = _normalize_score_sign_local(inst["name"], [nr_score])[0]
        if norm_nr == 0:
            continue
        gap_pct = (norm_normal - norm_nr) / abs(norm_nr) * 100.0
        if gap_pct > 0:
            wins_normal += 1
            verdict = "normal better"
        elif gap_pct < 0:
            wins_no_repulsion += 1
            verdict = "no_repulsion better"
        else:
            verdict = "tie"
        gap_values.append(gap_pct)
        gap_lines.append(
            f"{inst_name}: normal={norm_normal:.6f}, no_repulsion={norm_nr:.6f}, gap={gap_pct:.2f}% ({verdict})"
        )

        _best_algo, _best_score, rank_normal, _n_rank, _pct, _my_cmp, _best_cmp = (
            _rank_vs_global_ranking_excluding_ppo(
                repo_root, inst["name"], inst["dim"], inst["type_instance"], normal_score
            )
        )
        if rank_normal is not None:
            ranks_normal.append(rank_normal)
        _best_algo, _best_score, rank_nr, _n_rank, _pct, _my_cmp, _best_cmp = (
            _rank_vs_global_ranking_excluding_ppo(
                repo_root, inst["name"], inst["dim"], inst["type_instance"], nr_score
            )
        )
        if rank_nr is not None:
            ranks_no_repulsion.append(rank_nr)

    summary_path = Path(config_dir) / "normal_vs_no_repulsion.txt"
    with open(summary_path, "w") as f:
        f.write("Normal vs No_Repulsion summary\n")
        f.write(f"wins_normal: {wins_normal}\n")
        f.write(f"wins_no_repulsion: {wins_no_repulsion}\n")
        mean_gap = sum(gap_values) / len(gap_values) if gap_values else 0.0
        f.write(f"mean_gap: {mean_gap:.2f}%\n")
        mean_rank_normal = float(np.mean(ranks_normal)) if ranks_normal else None
        median_rank_normal = float(np.median(ranks_normal)) if ranks_normal else None
        mean_rank_nr = float(np.mean(ranks_no_repulsion)) if ranks_no_repulsion else None
        median_rank_nr = float(np.median(ranks_no_repulsion)) if ranks_no_repulsion else None
        f.write(f"mean_rank_normal: {mean_rank_normal}\n")
        f.write(f"median_rank_normal: {median_rank_normal}\n")
        f.write(f"mean_rank_no_repulsion: {mean_rank_nr}\n")
        f.write(f"median_rank_no_repulsion: {median_rank_nr}\n")
        f.write("per_instance_gap:\n")
        for line in gap_lines:
            f.write(f"{line}\n")
    print(f"[DONE] Wrote summary to {summary_path}")

    print(f"[DONE] overall no_repulsion summary at {out_xlsx}")
    print(f"Elapsed: {time.time() - start_all:.2f}s")


if __name__ == "__main__":
    main()
