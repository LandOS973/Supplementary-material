#!/usr/bin/env python3
import argparse
import os
import re
from pathlib import Path

import pandas as pd
from openpyxl.chart import BarChart, Reference


GROUP_RE = re.compile(r"^(?P<problem>[A-Za-z0-9]+)_dim(?P<dim>\d+)_t(?P<type_instance>\d+)$")
DUMMY_RE = re.compile(r"^dummy(?P<dummy_blocks>\d+)$")


def _parse_summary_config(summary_path):
    cfg = {}
    try:
        with open(summary_path, "r") as f:
            for raw_line in f:
                line = raw_line.strip()
                if not line or ":" not in line:
                    continue
                key, value = line.split(":", 1)
                key = key.strip().lower()
                value = value.strip()
                if not value:
                    continue
                lowered = value.lower()
                if lowered in ("none", "null", "n/a"):
                    parsed = None
                elif lowered in ("true", "false"):
                    parsed = lowered == "true"
                else:
                    try:
                        # handle scientific notation like 1e-4
                        if any(token in value for token in (".", "e", "E")):
                            parsed = float(value)
                        else:
                            parsed = int(value)
                    except ValueError:
                        parsed = value
                cfg[key] = parsed
    except OSError:
        return None
    return cfg


def _is_maximization_problem(problem_type: str) -> bool:
    return str(problem_type).upper() in ("NK", "BLOCK")


def _round_float(value, ndigits: int = 8):
    try:
        return round(float(value), ndigits)
    except Exception:
        return value


def _format_float(value, ndigits: int):
    try:
        return f"{float(value):.{ndigits}f}"
    except Exception:
        return str(value)


def _build_hist(df, col, label_ndigits=None):
    base = df.dropna(subset=[col])
    counts = (
        base.groupby(col)
        .size()
        .reset_index(name="count")
        .sort_values(by=col)
    )
    den = len(base)
    if den:
        counts["%"] = counts["count"] / den * 100.0
    label_col = None
    if label_ndigits is not None:
        label_col = f"{col}_label"
        counts[label_col] = counts[col].apply(lambda v: _format_float(v, label_ndigits))
        cols = [label_col, col, "count"]
        if den:
            cols.append("%")
        counts = counts[cols]
    return counts, label_col


def _add_bar_chart(sheet, df, title, category_col, value_col="count"):
    if df.empty or sheet.max_row < 2:
        return
    cols = list(df.columns)
    cat_idx = cols.index(category_col) + 1
    val_idx = cols.index(value_col) + 1
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = "count"
    data = Reference(sheet, min_col=val_idx, min_row=1, max_row=sheet.max_row)
    cats = Reference(sheet, min_col=cat_idx, min_row=2, max_row=sheet.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    sheet.add_chart(chart, "E2")


def _parse_group_from_path(path: Path):
    problem = None
    dim = None
    type_instance = None
    dummy_blocks = 0
    mode = "standard"
    no_interact = False

    parts = list(path.parts)
    for part in parts:
        match = GROUP_RE.match(part)
        if match:
            problem = match.group("problem")
            dim = int(match.group("dim"))
            type_instance = int(match.group("type_instance"))
        dummy_match = DUMMY_RE.match(part)
        if dummy_match:
            dummy_blocks = int(dummy_match.group("dummy_blocks"))
        if part == "no_interact":
            no_interact = True
        if part == "decay":
            mode = "decay"
    if problem is None or dim is None or type_instance is None:
        return None
    return {
        "problem": problem,
        "dim": dim,
        "type_instance": type_instance,
        "dummy_blocks": dummy_blocks,
        "no_interact": no_interact,
        "mode": mode,
    }


def _is_better(problem_type: str, score: float, best_score: float) -> bool:
    if best_score is None:
        return True
    if _is_maximization_problem(problem_type):
        return score > best_score
    return score < best_score


def main():
    parser = argparse.ArgumentParser(description="Review kernels and export results to Excel.")
    parser.add_argument(
        "--summary_root",
        type=str,
        default=None,
        help="Root folder containing *_best_summary.txt files (default: results/experiments).",
    )
    parser.add_argument(
        "--out",
        type=str,
        default=None,
        help="Output Excel path (default: courbes/review_kernels.xlsx).",
    )
    args = parser.parse_args()

    repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    summary_root = args.summary_root or os.path.join(repo_root, "results", "experiments")
    out_path = args.out or os.path.join(repo_root, "courbes", "review_kernels.xlsx")

    summary_root_path = Path(summary_root)
    summary_files = list(summary_root_path.rglob("*_best_summary.txt"))

    all_rows = []
    skipped = 0
    for path in summary_files:
        group = _parse_group_from_path(path)
        cfg = _parse_summary_config(path)
        if not group or not cfg:
            skipped += 1
            continue
        if str(group.get("problem", "")).upper() == "BLOCK":
            skipped += 1
            continue
        if group.get("no_interact"):
            skipped += 1
            continue
        if group.get("mode") not in ("standard", "decay"):
            skipped += 1
            continue
        kernel = cfg.get("kernel")
        avg_score = cfg.get("avg_score")
        if kernel is None or avg_score is None:
            skipped += 1
            continue
        row = {
            **group,
            "kernel": str(kernel).lower(),
            "avg_score": float(avg_score),
            "summary_path": str(path),
        }
        if "m" in cfg:
            row["M"] = cfg.get("m")
        if "lambda" in cfg:
            row["lambda"] = cfg.get("lambda")
        if "epsilon_svgd" in cfg:
            row["epsilon_svgd"] = _round_float(cfg.get("epsilon_svgd"))
        if "gamma" in cfg:
            row["gamma"] = _round_float(cfg.get("gamma"))
        if "decay_start_ratio" in cfg:
            row["decay_start_ratio"] = _round_float(cfg.get("decay_start_ratio"))
        if "decay_min_factor" in cfg:
            row["decay_min_factor"] = _round_float(cfg.get("decay_min_factor"))
        all_rows.append(row)

    if not all_rows:
        raise SystemExit(f"Aucun resume trouve dans {summary_root}.")

    all_df = pd.DataFrame(all_rows)
    all_df = all_df.drop(columns=["bandwith_kernel", "advantage"], errors="ignore")
    all_df = all_df.drop(columns=["dummy_blocks", "no_interact"], errors="ignore")

    standard_df = all_df[all_df["mode"] == "standard"].copy()
    decay_df = all_df[all_df["mode"] == "decay"].copy()

    best_rows = {}
    for _, row in standard_df.iterrows():
        key = (
            row["problem"],
            int(row["dim"]),
            int(row["type_instance"]),
            int(row.get("dummy_blocks", 0)),
            bool(row.get("no_interact", False)),
            row.get("mode", "standard"),
        )
        current = best_rows.get(key)
        if current is None:
            best_rows[key] = row
            continue
        if _is_better(row["problem"], row["avg_score"], current["avg_score"]):
            best_rows[key] = row

    best_df = pd.DataFrame(best_rows.values())
    best_df = best_df.drop(columns=["dummy_blocks", "no_interact"], errors="ignore")
    total_instances = len(best_df)

    kernel_counts, _ = _build_hist(best_df, "kernel")
    kernel_counts["%"] = kernel_counts["count"] / total_instances * 100.0 if total_instances else 0.0
    m_counts, _ = _build_hist(best_df, "M")
    lambda_counts, _ = _build_hist(best_df, "lambda")
    epsilon_counts, epsilon_label = _build_hist(best_df, "epsilon_svgd", label_ndigits=4)
    gamma_counts, gamma_label = _build_hist(best_df, "gamma", label_ndigits=5)

    decay_best_rows = {}
    for _, row in decay_df.iterrows():
        key = (
            row["problem"],
            int(row["dim"]),
            int(row["type_instance"]),
            int(row.get("dummy_blocks", 0)),
            bool(row.get("no_interact", False)),
            row.get("mode", "decay"),
        )
        current = decay_best_rows.get(key)
        if current is None:
            decay_best_rows[key] = row
            continue
        if _is_better(row["problem"], row["avg_score"], current["avg_score"]):
            decay_best_rows[key] = row

    decay_best_df = pd.DataFrame(decay_best_rows.values())
    if not decay_best_df.empty:
        decay_best_df = decay_best_df.drop(columns=["dummy_blocks", "no_interact"], errors="ignore")
        decay_instances = len(decay_best_df)
        decay_kernel_counts, _ = _build_hist(decay_best_df, "kernel")
        if decay_instances:
            decay_kernel_counts["%"] = decay_kernel_counts["count"] / decay_instances * 100.0
        decay_m_counts, _ = _build_hist(decay_best_df, "M")
        decay_lambda_counts, _ = _build_hist(decay_best_df, "lambda")
        decay_epsilon_counts, decay_epsilon_label = _build_hist(decay_best_df, "epsilon_svgd", label_ndigits=4)
        decay_gamma_counts, decay_gamma_label = _build_hist(decay_best_df, "gamma", label_ndigits=5)
        decay_start_counts, decay_start_label = _build_hist(decay_best_df, "decay_start_ratio", label_ndigits=4)
        decay_min_counts, decay_min_label = _build_hist(decay_best_df, "decay_min_factor", label_ndigits=4)
    else:
        decay_instances = 0
        decay_kernel_counts = pd.DataFrame()
        decay_m_counts = pd.DataFrame()
        decay_lambda_counts = pd.DataFrame()
        decay_epsilon_counts = pd.DataFrame()
        decay_gamma_counts = pd.DataFrame()
        decay_start_counts = pd.DataFrame()
        decay_min_counts = pd.DataFrame()
        decay_epsilon_label = None
        decay_gamma_label = None
        decay_start_label = None
        decay_min_label = None

    out_dir = os.path.dirname(out_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        best_df.sort_values(by=["problem", "dim", "type_instance", "kernel"]).to_excel(
            writer, sheet_name="best_by_instance", index=False
        )
        kernel_counts.to_excel(writer, sheet_name="kernel_share", index=False)
        standard_df.sort_values(by=["problem", "dim", "type_instance", "kernel"]).to_excel(
            writer, sheet_name="all_kernels", index=False
        )
        m_counts.to_excel(writer, sheet_name="M_hist", index=False)
        lambda_counts.to_excel(writer, sheet_name="lambda_hist", index=False)
        epsilon_counts.to_excel(writer, sheet_name="epsilon_hist", index=False)
        gamma_counts.to_excel(writer, sheet_name="gamma_hist", index=False)
        if not decay_best_df.empty:
            decay_best_df.sort_values(by=["problem", "dim", "type_instance", "kernel"]).to_excel(
                writer, sheet_name="decay_best_by_instance", index=False
            )
            decay_kernel_counts.to_excel(writer, sheet_name="decay_kernel_share", index=False)
            decay_m_counts.to_excel(writer, sheet_name="decay_M_hist", index=False)
            decay_lambda_counts.to_excel(writer, sheet_name="decay_lambda_hist", index=False)
            decay_epsilon_counts.to_excel(writer, sheet_name="decay_epsilon_hist", index=False)
            decay_gamma_counts.to_excel(writer, sheet_name="decay_gamma_hist", index=False)
            decay_start_counts.to_excel(writer, sheet_name="decay_start_hist", index=False)
            decay_min_counts.to_excel(writer, sheet_name="decay_min_hist", index=False)

        workbook = writer.book

        kernel_sheet = writer.sheets["kernel_share"]
        _add_bar_chart(kernel_sheet, kernel_counts, "Kernel (count)", "kernel")

        m_sheet = writer.sheets["M_hist"]
        _add_bar_chart(m_sheet, m_counts, "M (count)", "M")

        lambda_sheet = writer.sheets["lambda_hist"]
        _add_bar_chart(lambda_sheet, lambda_counts, "Lambda (count)", "lambda")

        epsilon_sheet = writer.sheets["epsilon_hist"]
        _add_bar_chart(
            epsilon_sheet,
            epsilon_counts,
            "Epsilon SVGD (count)",
            epsilon_label or "epsilon_svgd",
        )

        gamma_sheet = writer.sheets["gamma_hist"]
        _add_bar_chart(
            gamma_sheet,
            gamma_counts,
            "Gamma (count)",
            gamma_label or "gamma",
        )

        if not decay_best_df.empty:
            decay_kernel_sheet = writer.sheets["decay_kernel_share"]
            _add_bar_chart(decay_kernel_sheet, decay_kernel_counts, "Decay kernel (count)", "kernel")

            decay_m_sheet = writer.sheets["decay_M_hist"]
            _add_bar_chart(decay_m_sheet, decay_m_counts, "Decay M (count)", "M")

            decay_lambda_sheet = writer.sheets["decay_lambda_hist"]
            _add_bar_chart(decay_lambda_sheet, decay_lambda_counts, "Decay lambda (count)", "lambda")

            decay_epsilon_sheet = writer.sheets["decay_epsilon_hist"]
            _add_bar_chart(
                decay_epsilon_sheet,
                decay_epsilon_counts,
                "Decay epsilon (count)",
                decay_epsilon_label or "epsilon_svgd",
            )

            decay_gamma_sheet = writer.sheets["decay_gamma_hist"]
            _add_bar_chart(
                decay_gamma_sheet,
                decay_gamma_counts,
                "Decay gamma (count)",
                decay_gamma_label or "gamma",
            )

            decay_start_sheet = writer.sheets["decay_start_hist"]
            _add_bar_chart(
                decay_start_sheet,
                decay_start_counts,
                "Decay start ratio (count)",
                decay_start_label or "decay_start_ratio",
            )

            decay_min_sheet = writer.sheets["decay_min_hist"]
            _add_bar_chart(
                decay_min_sheet,
                decay_min_counts,
                "Decay min factor (count)",
                decay_min_label or "decay_min_factor",
            )

    print(f"Ecrit: {out_path}")
    print(f"Instances traitees: {total_instances}")
    if skipped:
        print(f"Resumes ignores: {skipped}")


if __name__ == "__main__":
    main()
