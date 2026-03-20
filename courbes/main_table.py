#!/usr/bin/env python3
"""Generate a summary table image across all tested instances."""

from __future__ import annotations

import argparse
import csv
import math
import re
import textwrap
from pathlib import Path
from typing import List, Optional, Sequence

import matplotlib.pyplot as plt


ROOT = Path(__file__).resolve().parent.parent
COMPETITOR_DIRS = [
    Path("/home/landos/Downloads/results_nevergrad_ppsn"),
]
DEFAULT_BUDGET = 50000


def parse_summary_file(path: Path) -> dict[str, str]:
    data: dict[str, str] = {}
    with path.open() as handle:
        for line in handle:
            if ":" not in line:
                continue
            key, value = line.split(":", 1)
            data[key.strip()] = value.strip()
    return data


def parse_float(value: str | None) -> float | None:
    if value is None:
        return None
    try:
        parsed = float(value)
        if not math.isfinite(parsed):
            return None
        return parsed
    except ValueError:
        return None


def is_maximization_problem(problem_name: str) -> bool:
    return problem_name.upper() in ("NK", "NK3", "BLOCK")


def normalize_score(problem_name: str, value: float | None) -> float | None:
    if value is None:
        return None
    if is_maximization_problem(problem_name):
        return value
    return -value if value < 0 else value


def parse_rank_value(value: str | None) -> tuple[int, int] | None:
    if not value:
        return None
    match = re.search(r"(\d+)\s*/\s*(\d+)", value)
    if not match:
        return None
    return int(match.group(1)), int(match.group(2))


def _normalize_scores(problem_name: str, scores: Sequence[float]) -> List[float]:
    normalized: List[float] = []
    for value in scores:
        norm = normalize_score(problem_name, value)
        if norm is None:
            continue
        normalized.append(norm)
    return normalized


def load_raw_scores_csv(path: Path) -> List[float]:
    if not path.exists():
        return []
    scores: List[float] = []
    try:
        with path.open(newline="") as handle:
            reader = csv.reader(handle)
            header = next(reader, None)
            if header is None:
                return []
            header_lower = [h.strip().lower() for h in header]
            if "score" in header_lower:
                idx = header_lower.index("score")
                for row in reader:
                    if len(row) <= idx:
                        continue
                    value = parse_float(row[idx])
                    if value is not None:
                        scores.append(value)
            else:
                # Header is not explicit; attempt to parse header as data too.
                for row in [header] + list(reader):
                    if not row:
                        continue
                    value = parse_float(row[0])
                    if value is not None:
                        scores.append(value)
    except Exception:
        return []
    return scores


def read_last_numeric_score(path: Path) -> float | None:
    try:
        lines = path.read_text().splitlines()
    except OSError:
        return None
    for line in reversed(lines):
        line = line.strip()
        if not line:
            continue
        if line.lower().startswith("runtime"):
            continue
        parts = [part.strip() for part in line.split(",") if part.strip()]
        for part in reversed(parts):
            try:
                return float(part)
            except ValueError:
                continue
    return None


def _problem_variants(problem_name: str | None) -> List[str]:
    if not problem_name:
        return ["UBQP", "QUBO", "qubo", "ubqp"]
    normalized = problem_name.upper()
    if normalized in ("QUBO", "UBQP"):
        return ["UBQP", "QUBO", "qubo", "ubqp"]
    return [normalized, normalized.lower()]


def find_competitor_run_files(
    algo: str,
    dim: int,
    type_instance: int,
    problem_name: str | None = None,
    budget: int = DEFAULT_BUDGET,
) -> List[Path]:
    candidate_files: List[Path] = []
    problem_variants = _problem_variants(problem_name)
    for root in COMPETITOR_DIRS:
        algo_dir = root / algo
        if not algo_dir.exists():
            continue
        for problem in problem_variants:
            candidate_dir = algo_dir / problem / str(dim) / str(type_instance)
            if not candidate_dir.exists():
                continue
            files = sorted(candidate_dir.glob(f"*_budget_{budget}_*.txt"))
            if not files:
                files = sorted(candidate_dir.glob("*.txt"))
            candidate_files.extend(files)
        if candidate_files:
            break

        # Fallback: recursive search under algo dir.
        patterns: List[str] = []
        for problem in problem_variants:
            patterns.extend(
                (
                    f"**/results_nevergrad_{algo}_{problem}_{dim}_{type_instance}_*.txt",
                    f"**/*{algo}*{problem}*{dim}*{type_instance}*.txt",
                )
            )
        for pattern in patterns:
            candidate_files.extend(algo_dir.glob(pattern))
        if candidate_files:
            break

    return sorted(set(candidate_files))


def load_competitor_final_scores(
    algo: str,
    dim: int,
    type_instance: int,
    problem_name: str | None = None,
) -> List[float]:
    paths = find_competitor_run_files(algo, dim, type_instance, problem_name)
    scores: List[float] = []
    for path in paths:
        value = read_last_numeric_score(path)
        if value is not None:
            scores.append(value)
    return scores


def load_best_summary(experiment_dir: Path, problem_name: str) -> dict[str, str] | None:
    summary_files = sorted(experiment_dir.glob(f"{problem_name}_*_best_summary.txt"))
    if not summary_files:
        return None
    best_summary = None
    best_score = None
    maximize = is_maximization_problem(problem_name)
    for path in summary_files:
        data = parse_summary_file(path)
        avg_score = parse_float(data.get("avg_score"))
        if avg_score is None:
            continue
        if best_score is None:
            best_score = avg_score
            best_summary = data
            continue
        if maximize and avg_score > best_score:
            best_score = avg_score
            best_summary = data
        elif not maximize and avg_score < best_score:
            best_score = avg_score
            best_summary = data
    return best_summary


def load_last_metrics_score(csv_path: Path) -> float | None:
    """Load the score from the last data row of a best_metrics.csv file."""
    if not csv_path.exists():
        return None

    def _data_lines(handle):
        for line in handle:
            if not line.strip():
                continue
            if line.lstrip().startswith("#"):
                continue
            yield line

    try:
        with csv_path.open() as handle:
            reader = csv.reader(_data_lines(handle))
            header = next(reader, None)
            if not header:
                return None
            last_row = None
            for row in reader:
                last_row = row
            if not last_row:
                return None
            row_dict = dict(zip(header, last_row))
            for key in ("mean", "avg_score", "score", "avg", "avg_fitness", "fitness", "best_fitness"):
                if key in row_dict and row_dict[key]:
                    return parse_float(row_dict[key])
    except Exception:
        return None
    return None


def load_svgd_score(config_dir: Path, problem_name: str, dim: int, type_instance: int) -> tuple[float | None, tuple[int, int] | None]:
    """Load SVGD score and optional rank from config directory.

    Returns (score, (rank, total)) where rank tuple may be None.
    """
    instance_dir = config_dir / f"{problem_name}_dim{dim}_t{type_instance}"
    if not instance_dir.exists():
        return None, None

    csv_path = instance_dir / "best_metrics.csv"
    csv_score = load_last_metrics_score(csv_path)
    if csv_score is None:
        return None, None
    return csv_score, None


def ranking_path(problem_name: str, dim: int, type_instance: int) -> Path:
    if problem_name.upper() == "QUBO":
        name = "UBQP"
    else:
        name = problem_name
    return ROOT / "additional_results" / "global_ranking" / f"{name}_N_{dim}_K_{type_instance}_ranks.csv"


def load_ranking(path: Path) -> List[tuple[str, float]]:
    if not path.exists():
        return []
    rows: List[tuple[str, float]] = []
    with path.open(newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            name = (row.get("name_algo") or row.get("name") or row.get("algo") or "").strip()
            score = parse_float(row.get("score"))
            if not name or score is None:
                continue
            rows.append((name, score))
    return rows


def discover_instances() -> List[tuple[str, int, int, Optional[Path]]]:
    exp_root = ROOT / "results" / "experiments"
    if not exp_root.exists():
        return []
    instances: List[tuple[str, int, int, Optional[Path]]] = []
    seen: set[tuple[str, int, int]] = set()
    for entry in exp_root.iterdir():
        if not entry.is_dir():
            continue
        match = re.match(r"^(?P<name>.+)_dim(?P<dim>\d+)_t(?P<t>\d+)$", entry.name)
        if not match:
            continue
        name = match.group("name")
        dim = int(match.group("dim"))
        type_instance = int(match.group("t"))
        key = (name, dim, type_instance)
        if key in seen:
            continue
        seen.add(key)
        instances.append((name, dim, type_instance, entry))
    return sorted(instances, key=lambda item: (item[0], item[1], item[2]))


def discover_instances_from_rankings() -> List[tuple[str, int, int, Optional[Path]]]:
    ranking_dir = ROOT / "additional_results" / "global_ranking"
    if not ranking_dir.exists():
        return []
    instances: List[tuple[str, int, int, Optional[Path]]] = []
    for rank_file in ranking_dir.glob("*_ranks.csv"):
        match = re.match(r"^(?P<problem>UBQP|NK|NK3)_N_(?P<dim>\d+)_K_(?P<t>\d+)_ranks\.csv$", rank_file.name)
        if not match:
            continue
        if match.group("problem") == "UBQP":
            problem = "QUBO"
        else:
            problem = match.group("problem")
        dim = int(match.group("dim"))
        type_instance = int(match.group("t"))
        instances.append((problem, dim, type_instance, None))
    return instances


def format_rank(rank: int | None, total: int | None) -> str:
    if rank is None or total is None:
        return "—"
    return f"{rank}/{total}"


def format_score(problem_name: str, value: float | None) -> str:
    if value is None:
        return "—"
    value = abs(value)
    if problem_name.upper() in ("NK", "NK3"):
        return f"{value:.4f}"
    return f"{value:.1f}"


def wrap_name(name: str | None, width: int = 18) -> str:
    if not name:
        return "—"
    return textwrap.fill(name, width=width)


def build_rows(
    methods: List[str], config_dir: Path
) -> tuple[
    List[List[str]],
    List[str],
    dict[tuple[str, int, int], float | None],
    dict[tuple[str, int, int], str | None],
]:
    rows: List[List[str]] = []
    raw_svgd_scores: dict[tuple[str, int, int], float | None] = {}
    best_other_map: dict[tuple[str, int, int], str | None] = {}

    # Gather instances from results/experiments
    instances = list(discover_instances())

    # Also gather instances present in the config directory (to ensure none missing)
    if config_dir.exists():
        for entry in config_dir.iterdir():
            # looking for directories like NK_dim64_t1
            match = re.match(r"^(?P<name>.+)_dim(?P<dim>\d+)_t(?P<t>\d+)$", entry.name)
            if not match:
                continue
            name = match.group("name")
            dim = int(match.group("dim"))
            type_instance = int(match.group("t"))
            key = (name, dim, type_instance)
            if key not in {(i[0], i[1], i[2]) for i in instances}:
                instances.append((name, dim, type_instance, entry))

    # Also gather instances present in global rankings (NK3, NK, QUBO)
    ranking_instances = discover_instances_from_rankings()
    existing_keys = {(i[0], i[1], i[2]) for i in instances}
    for name, dim, type_instance, _ in ranking_instances:
        key = (name, dim, type_instance)
        if key not in existing_keys:
            instances.append((name, dim, type_instance, None))
            existing_keys.add(key)

    # sort instances for deterministic output
    instances = sorted(instances, key=lambda item: (item[0], item[1], item[2]))

    for problem_name, dim, type_instance, exp_dir in instances:
        # Exclude specific problematic instance per user request
        if problem_name.upper() == "BLOCK" and dim == 2064 and type_instance == 16:
            continue
        # Try to load summary from experiments folder if exists
        avg_score = None
        exp_summary = None
        if exp_dir and exp_dir.exists():
            exp_summary = load_best_summary(exp_dir, problem_name)
            if exp_summary:
                avg_score = parse_float(exp_summary.get("avg_score"))

        row: List[str] = [problem_name, str(dim), str(type_instance)]

        # Load SVGD score and optional rank from config
        svgd_raw_score, svgd_rank_from_config = load_svgd_score(config_dir, problem_name, dim, type_instance)
        svgd_score = normalize_score(problem_name, svgd_raw_score)
        raw_svgd_scores[(problem_name, dim, type_instance)] = svgd_score

        ranking_file = ranking_path(problem_name, dim, type_instance)
        ranking = load_ranking(ranking_file)
        ranking = [
            (name, score)
            for name, score in ranking
            if name not in {"PPO-EDA", "Tabu", "TABU"}
        ]
        ranking = sorted(ranking, key=lambda item: item[1], reverse=True)
        combined = list(ranking)
        if svgd_score is not None:
            combined.append(("SVGD", svgd_score))
        combined.sort(key=lambda item: item[1], reverse=True)
        total_competitors = len(ranking)
        total = len(combined) if combined else None
        # Ranking including SVGD (SVGD participates in the ranking)
        rank_map: dict[str, tuple[int, float]] = {}
        for idx, (name, score) in enumerate(combined, start=1):
            rank_map[name] = (idx, score)

        svgd_rank = rank_map.get("SVGD", (None, None))[0]
        row.extend([format_rank(svgd_rank, total), format_score(problem_name, svgd_score)])

        # Add other methods
        for method in methods:
            rank_entry = rank_map.get(method)
            if rank_entry:
                rank, score = rank_entry
                row.extend([format_rank(rank, total), format_score(problem_name, score)])
            else:
                row.extend(["—", "—"])

        # Best method (excluding SVGD, PBIL, MIMIC, BOA)
        best_name = "—"
        best_rank = None
        best_score = None
        if combined:
            excluded = set(methods)
            excluded.add("SVGD")
            for idx, (name, score) in enumerate(combined, start=1):
                if name in excluded:
                    continue
                best_name = name
                best_rank = idx
                best_score = score
                break
        best_other_map[(problem_name, dim, type_instance)] = None if best_name == "—" else best_name

        row.extend(
            [
                wrap_name(best_name),
                format_rank(best_rank, total),
                format_score(problem_name, best_score),
            ]
        )
        rows.append(row)

    return (
        rows,
        ["Pb", "n", "t"] + ["Rank", "Score"] * (1 + len(methods)) + ["Name", "Rank", "Score"],
        raw_svgd_scores,
        best_other_map,
    )


def _latex_escape(text: str) -> str:
    return (
        text.replace("\\", "\\textbackslash{}")
        .replace("_", "\\_")
        .replace("&", "\\&")
        .replace("%", "\\%")
        .replace("#", "\\#")
    )


def _parse_rank_value(rank_str: str | None) -> int | None:
    if not rank_str or rank_str == "—":
        return None
    parsed = parse_rank_value(rank_str)
    if not parsed:
        return None
    return parsed[0]


def _parse_score_value(score_str: str | None) -> float | None:
    if not score_str or score_str == "—":
        return None
    # Accept optional LaTeX decorations such as $^{*}$.
    cleaned = score_str.replace("$^{*}$", "").replace("\\textsuperscript{*}", "")
    match = re.search(r"-?\d+(?:\.\d+)?(?:[eE][+-]?\d+)?", cleaned)
    if not match:
        return None
    try:
        return float(match.group(0))
    except Exception:
        return None


def _best_wrap(text: str) -> str:
    return f"\\best{{{text}}}"

def write_latex_table(
    rows: List[List[str]],
    output_tex: Path,
    raw_svgd_scores: dict[tuple[str, int, int], float | None] | None = None,
    config_dir: Path | None = None,
    best_other_map: dict[tuple[str, int, int], str | None] | None = None,
) -> None:
    if not rows:
        print("[WARN] No rows to write.")
        return

    try:
        from scipy import stats as scipy_stats
    except Exception:
        scipy_stats = None

    warned: set[str] = set()
    svgd_scores_cache: dict[tuple[str, int, int], List[float]] = {}
    competitor_scores_cache: dict[tuple[str, str, int, int], List[float]] = {}

    def warn_once(key: str, message: str) -> None:
        if key in warned:
            return
        warned.add(key)
        print(f"[WARN] {message}")

    def _finite(values: Sequence[float]) -> List[float]:
        return [v for v in values if math.isfinite(v)]

    def _load_svgd_scores(problem_name: str, dim: int, type_instance: int) -> List[float]:
        key = (problem_name, dim, type_instance)
        if key in svgd_scores_cache:
            return svgd_scores_cache[key]
        if config_dir is None:
            return []
        raw_path = config_dir / f"{problem_name}_dim{dim}_t{type_instance}" / "raw_scores.csv"
        scores = _normalize_scores(problem_name, load_raw_scores_csv(raw_path))
        scores = _finite(scores)
        if not scores:
            warn_once(
                f"svgd_missing_{problem_name}_{dim}_{type_instance}",
                f"Missing or empty raw_scores.csv for SVGD ({problem_name} dim={dim} t={type_instance}).",
            )
        svgd_scores_cache[key] = scores
        return scores

    def _load_competitor_scores(
        algo: str, problem_name: str, dim: int, type_instance: int
    ) -> List[float]:
        key = (algo, problem_name, dim, type_instance)
        if key in competitor_scores_cache:
            return competitor_scores_cache[key]
        scores = _normalize_scores(
            problem_name, load_competitor_final_scores(algo, dim, type_instance, problem_name)
        )
        scores = _finite(scores)
        if not scores:
            warn_once(
                f"comp_missing_{algo}_{problem_name}_{dim}_{type_instance}",
                f"Missing competitor scores for {algo} ({problem_name} dim={dim} t={type_instance}).",
            )
        competitor_scores_cache[key] = scores
        return scores

    def _star_target(
        problem_name: str,
        dim: int,
        type_instance: int,
        best_algo: str | None,
        svgd_display: float | None,
        best_display: float | None,
    ) -> str | None:
        if best_algo is None:
            return None
        if scipy_stats is None:
            warn_once("scipy_missing", "scipy is not installed; skipping t-tests.")
            return None
        svgd_scores = _load_svgd_scores(problem_name, dim, type_instance)
        competitor_scores = _load_competitor_scores(best_algo, problem_name, dim, type_instance)
        if len(svgd_scores) < 2 or len(competitor_scores) < 2:
            warn_once(
                f"ttest_missing_{problem_name}_{dim}_{type_instance}",
                f"Insufficient scores for t-test ({problem_name} dim={dim} t={type_instance}).",
            )
            return None
        result = scipy_stats.ttest_ind(
            svgd_scores, competitor_scores, equal_var=True, nan_policy="omit"
        )
        pvalue = getattr(result, "pvalue", None)
        if pvalue is None or not math.isfinite(pvalue):
            return None
        if pvalue >= 0.001:
            return None
        if svgd_display is None or best_display is None:
            return None
        # Use the displayed (aggregated) values to decide which score gets the star.
        if svgd_display > best_display:
            return "svgd"
        if best_display > svgd_display:
            return "best"
        return None

    def _pvalue_against(
        algo: str | None,
        problem_name: str,
        dim: int,
        type_instance: int,
    ) -> float | None:
        if algo is None:
            return None
        if scipy_stats is None:
            return None
        svgd_scores = _load_svgd_scores(problem_name, dim, type_instance)
        competitor_scores = _load_competitor_scores(algo, problem_name, dim, type_instance)
        if len(svgd_scores) < 2 or len(competitor_scores) < 2:
            return None
        result = scipy_stats.ttest_ind(
            svgd_scores, competitor_scores, equal_var=True, nan_policy="omit"
        )
        pvalue = getattr(result, "pvalue", None)
        if pvalue is None or not math.isfinite(pvalue):
            return None
        return float(pvalue)

    output_tex.parent.mkdir(parents=True, exist_ok=True)
    with output_tex.open("w") as f:
        f.write("% Table: Global rankings and average scores\n")
        f.write("\\begin{table}[htbp]\n")
        f.write("    \\centering\n")
        f.write("    \\resizebox{\\textwidth}{!}{%\n")
        f.write("    \\begin{tabular}{ccc cc cc cc cc lcc}\n")
        f.write("        \\toprule\n")
        f.write("        \\multicolumn{3}{c}{\\textbf{Instances}} & \n")
        f.write("        \\multicolumn{2}{c}{\\textbf{\\texttt{SVGD-EDA}}} & \n")
        f.write("        \\multicolumn{2}{c}{\\textbf{\\texttt{PBIL}}} & \n")
        f.write("        \\multicolumn{2}{c}{\\textbf{\\texttt{MIMIC}}} & \n")
        f.write("        \\multicolumn{2}{c}{\\textbf{\\texttt{BOA}}} & \n")
        f.write("        \\multicolumn{3}{c}{\\textbf{Best Method (Others)}} \\\\\n")
        f.write("        \\cmidrule(r){1-3} \\cmidrule(lr){4-5} \\cmidrule(lr){6-7} ")
        f.write("\\cmidrule(lr){8-9} \\cmidrule(lr){10-11} \\cmidrule(l){12-14}\n")
        f.write("        Pb & $n$ & $t$ & Rank & Score & Rank & Score & Rank & Score & Rank & Score & Name & Rank & Score \\\\\n")
        f.write("        \\midrule\n")

        # Compute mean ranks and mean relative scores (score / best score per row)
        rank_indices = [3, 5, 7, 9, 12]
        score_indices = [4, 6, 8, 10, 13]
        mean_ranks = [[] for _ in rank_indices]
        mean_rel_scores = [[] for _ in score_indices]

        # Find the most frequent "best method (others)" across rows.
        best_method_counts: dict[str, int] = {}
        for row in rows:
            if len(row) <= 11:
                continue
            name_raw = row[11]
            if not name_raw or name_raw == "—":
                continue
            name_flat = re.sub(r"\s+", "", name_raw)
            if not name_flat:
                continue
            best_method_counts[name_flat] = best_method_counts.get(name_flat, 0) + 1
        most_common_best = None
        if best_method_counts:
            most_common_best = max(best_method_counts.items(), key=lambda item: item[1])[0]

        for row in rows:
            scores = [_parse_score_value(row[i]) for i in score_indices]
            if raw_svgd_scores is not None:
                try:
                    key = (row[0], int(row[1]), int(row[2]))
                except Exception:
                    key = None
                if key is not None:
                    svgd_raw = raw_svgd_scores.get(key)
                    if svgd_raw is not None:
                        scores[0] = svgd_raw
            score_vals = [v for v in scores if v is not None]
            best_score = max(score_vals) if score_vals else None
            for idx, col in enumerate(rank_indices):
                r = _parse_rank_value(row[col])
                if r is not None:
                    mean_ranks[idx].append(float(r))
            if best_score:
                for idx, s in enumerate(scores):
                    if s is not None:
                        mean_rel_scores[idx].append(s / best_score)

        def fmt_mean(values, ndigits):
            if not values:
                return "—"
            return f"{sum(values) / len(values):.{ndigits}f}"

        # Mean rank/score for the most frequent "best method (others)"
        best_method_ranks = []
        best_method_rel_scores = []
        if most_common_best:
            for row in rows:
                if len(row) <= 13:
                    continue
                problem_name = row[0]
                try:
                    dim = int(row[1])
                    type_instance = int(row[2])
                except Exception:
                    continue

                ranking_file = ranking_path(problem_name, dim, type_instance)
                ranking = load_ranking(ranking_file)
                ranking = [
                    (name, score)
                    for name, score in ranking
                    if name not in {"PPO-EDA", "Tabu", "TABU"}
                ]
                ranking = sorted(ranking, key=lambda item: item[1], reverse=True)
                if not ranking:
                    continue

                if raw_svgd_scores is not None:
                    svgd_score = raw_svgd_scores.get((problem_name, dim, type_instance))
                else:
                    svgd_score = _parse_score_value(row[4]) if len(row) > 4 else None
                combined = list(ranking)
                if svgd_score is not None:
                    combined.append(("SVGD", svgd_score))
                combined.sort(key=lambda item: item[1], reverse=True)
                rank_map = {name: (idx + 1, score) for idx, (name, score) in enumerate(combined)}
                if most_common_best not in rank_map:
                    continue
                rank_val, score_val = rank_map[most_common_best]
                best_score = combined[0][1]
                best_method_ranks.append(float(rank_val))
                if best_score:
                    best_method_rel_scores.append(score_val / best_score)

        # Les valeurs formatées donneront exactement "7.670", "0.987", etc.
        mean_row = [
            "Global ranking",
            "", 
            "",
            fmt_mean(mean_ranks[0], 3),
            fmt_mean(mean_rel_scores[0], 3),
            fmt_mean(mean_ranks[1], 3),
            fmt_mean(mean_rel_scores[1], 3),
            fmt_mean(mean_ranks[2], 3),
            fmt_mean(mean_rel_scores[2], 3),
            fmt_mean(mean_ranks[3], 3),
            fmt_mean(mean_rel_scores[3], 3),
            most_common_best or "DiscreteDE",
            fmt_mean(best_method_ranks, 3),
            fmt_mean(best_method_rel_scores, 3),
        ]

        # --- Début de la logique de regroupement (multirow) ---
        problem_counts = {}
        dim_counts = {}
        for row in rows:
            pb = row[0]
            dim = row[1]
            problem_counts[pb] = problem_counts.get(pb, 0) + 1
            key_dim = (pb, dim)
            dim_counts[key_dim] = dim_counts.get(key_dim, 0) + 1

        pb_seen = {}
        dim_seen = {}
        last_pb = None
        last_dim = None

        for row_index, row in enumerate(rows):
            pb = row[0]
            dim = row[1]
            
            # Gestion des lignes de séparation avec \midrule[1.2pt] entre les problèmes
            if last_pb is not None and pb != last_pb:
                f.write("        \\midrule[1.2pt]\n")
            elif last_dim is not None and dim != last_dim and pb == last_pb:
                f.write("        \\cmidrule(lr){2-14}\n")

            rank_values = [v for v in (_parse_rank_value(row[i]) for i in rank_indices) if v is not None]
            score_values = [v for v in (_parse_score_value(row[i]) for i in score_indices) if v is not None]
            best_rank = min(rank_values) if rank_values else None
            best_score = max(score_values) if score_values else None

            star = None
            try:
                dim_int = int(dim)
                type_int = int(row[2])
            except Exception:
                dim_int = None
                type_int = None
            if dim_int is not None and type_int is not None:
                key = (pb, dim_int, type_int)
                best_algo = None
                if best_other_map is not None:
                    best_algo = best_other_map.get(key)
                if best_algo is None:
                    name_raw = row[11] if len(row) > 11 else None
                    if name_raw and name_raw != "—":
                        best_algo = re.sub(r"\s+", "", name_raw)
                svgd_display = _parse_score_value(row[4]) if len(row) > 4 else None
                best_display = _parse_score_value(row[13]) if len(row) > 13 else None
                star = _star_target(
                    pb,
                    dim_int,
                    type_int,
                    best_algo,
                    svgd_display,
                    best_display,
                )

                # Print p-values for this line (SVGD vs PBIL/MIMIC/BOA and best method).
                p_pbil = _pvalue_against("PBIL", pb, dim_int, type_int)
                p_mimic = _pvalue_against("MIMIC", pb, dim_int, type_int)
                p_boa = _pvalue_against("BOA", pb, dim_int, type_int)
                p_best = _pvalue_against(best_algo, pb, dim_int, type_int)
                print(
                    "PVAL",
                    f"{pb} dim={dim_int} t={type_int}",
                    "| SVGD vs PBIL:", "NA" if p_pbil is None else f"{p_pbil:.6g}",
                    "| SVGD vs MIMIC:", "NA" if p_mimic is None else f"{p_mimic:.6g}",
                    "| SVGD vs BOA:", "NA" if p_boa is None else f"{p_boa:.6g}",
                    "| SVGD vs", best_algo or "—", ":", "NA" if p_best is None else f"{p_best:.6g}",
                )

            row_display = list(row)
            if star == "svgd" and len(row_display) > 4 and row_display[4] != "—":
                row_display[4] = f"{row_display[4]}$^{{*}}$"
            elif star == "best" and len(row_display) > 13 and row_display[13] != "—":
                row_display[13] = f"{row_display[13]}$^{{*}}$"

            cells: List[str] = []
            for idx, val in enumerate(row_display):
                if idx == 0:
                    if pb_seen.get(pb, 0) == 0:
                        cell = f"\\multirow{{{problem_counts[pb]}}}{{*}}{{{_latex_escape(val)}}}"
                    else:
                        cell = ""
                    pb_seen[pb] = pb_seen.get(pb, 0) + 1
                elif idx == 1:
                    key_dim = (pb, dim)
                    if dim_seen.get(key_dim, 0) == 0:
                        cell = f"\\multirow{{{dim_counts[key_dim]}}}{{*}}{{{_latex_escape(val)}}}"
                    else:
                        cell = ""
                    dim_seen[key_dim] = dim_seen.get(key_dim, 0) + 1
                elif idx == 11:
                    if val == "—":
                        cell = "—"
                    else:
                        name_flat = re.sub(r"\s+", "", val)
                        cell = f"\\mbox{{\\texttt{{{_latex_escape(name_flat)}}}}}"
                else:
                    cell = _latex_escape(val)

                if idx in rank_indices and best_rank is not None:
                    rank_val = _parse_rank_value(val)
                    if rank_val == best_rank:
                        cell = _best_wrap(cell)
                if idx in score_indices and best_score is not None:
                    score_val = _parse_score_value(val)
                    if score_val is not None and abs(score_val - best_score) <= 1e-9:
                        cell = _best_wrap(cell)
                cells.append(cell)

            f.write("        " + " & ".join(cells) + " \\\\\n")
            last_pb = pb
            last_dim = dim
        # --- Fin de la logique de regroupement ---

        # Ecriture de la ligne finale (résumé formaté)
        f.write("        \\midrule\n")
        row = mean_row
        svgd_mean_rank = _latex_escape(row[3])
        svgd_mean_rel = _latex_escape(row[4])

        best_name_raw = row[11]
        if best_name_raw == "—":
            best_name_cell = "—"
        else:
            name_flat = re.sub(r"\s+", "", best_name_raw)
            best_name_cell = f"\\textbf{{{_latex_escape(name_flat)}}}"

        best_mean_rank = _latex_escape(row[12])
        best_mean_rel = _latex_escape(row[13])

        cells = [
            "\\multicolumn{3}{c}{\\textbf{SVGD-EDA}}",
            f"{svgd_mean_rank}",
            f"{svgd_mean_rel}",
            "\\multicolumn{2}{c}{}",
            "\\multicolumn{2}{c}{}",
            "\\multicolumn{2}{c}{}",
            best_name_cell,
            f"{best_mean_rank}",
            f"{best_mean_rel}",
        ]

        f.write("        " + " & ".join(cells) + " \\\\\n")

        f.write("        \\bottomrule\n")
        f.write("    \\end{tabular}%\n")
        f.write("    }\n")
        f.write(
            "    \\caption{Global rankings and average scores obtained by \\texttt{SVGD-EDA} and the other EDAs "
            "(\\texttt{PBIL}, \\texttt{MIMIC}, and \\texttt{BOA}) are reported. The last columns present the ranking "
            "and average score of the best-performing method among the additional algorithms considered. "
            "Rankings are computed by comparing the best score achieved after 50,000 objective function evaluations, "
            "averaged across 100 independent runs. The last row reports global mean ranks and mean relative scores "
            "(score normalized by the best score of each instance distribution) for \\texttt{SVGD-EDA} and for the "
            "most frequently top-ranked method among the other algorithms, with that method's rank and relative "
            "score averaged across all instances. Bold values highlight the best results among all competing methods.}\n"
        )
        f.write("    \\label{tab:results_portrait}\n")
        f.write("\\end{table}\n")
    print(f"Saved LaTeX table to {output_tex}")


def write_csv(rows: List[List[str]], col_labels: List[str], output_csv: Path) -> None:
    if not rows:
        print("[WARN] No rows to write.")
        return
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    with output_csv.open("w", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(col_labels)
        writer.writerows(rows)
    print(f"Saved table to {output_csv}")


def write_excel(rows: List[List[str]], col_labels: List[str], output_xlsx: Path) -> None:
    if not rows:
        print("[WARN] No rows to write.")
        return
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        print("[WARN] openpyxl is not installed. Install it or use --format csv.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"

    # Column groups (0-based indices)
    # Instances: Pb, n, t -> cols 1-3
    # Methods: SVGD, PBIL, MIMIC, BOA -> each has Rank/Score
    # Best method (others): Name, Rank, Score -> last 3
    n_cols = len(col_labels)
    if n_cols != 14:
        print(f"[WARN] Unexpected column count: {n_cols}, expected 14. Excel header merging may be off.")

    # Header rows
    ws.append([""] * n_cols)
    ws.append([""] * n_cols)
    ws.append(col_labels)

    # Group headers (row 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws.cell(row=1, column=1, value="Instances")
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=11)
    ws.cell(row=1, column=4, value="Methods")
    ws.merge_cells(start_row=1, start_column=12, end_row=1, end_column=14)
    ws.cell(row=1, column=12, value="Best method (others)")

    # Method headers (row 2)
    ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=5)
    ws.cell(row=2, column=4, value="SVGD")
    ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=7)
    ws.cell(row=2, column=6, value="PBIL")
    ws.merge_cells(start_row=2, start_column=8, end_row=2, end_column=9)
    ws.cell(row=2, column=8, value="MIMIC")
    ws.merge_cells(start_row=2, start_column=10, end_row=2, end_column=11)
    ws.cell(row=2, column=10, value="BOA")

    # Data rows
    for row in rows:
        ws.append(row)

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx in range(1, 4):
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = bold
            cell.alignment = center

    for row_idx in range(4, 4 + len(rows)):
        for col_idx in range(1, n_cols + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = center

    col_widths = [8, 6, 6] + [9, 9] * 4 + [24, 9, 9]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    print(f"Saved table to {output_xlsx}")


def plot_table(rows: List[List[str]], col_labels: List[str], output_png: Path, output_pdf: Path) -> None:
    if not rows:
        print("[WARN] No rows to render.")
        return

    n_rows = len(rows)
    fig_height = max(8.0, 0.42 * (n_rows + 2))
    fig, ax = plt.subplots(figsize=(18, fig_height), dpi=220)
    ax.axis("off")
    ax.set_position([0.02, 0.02, 0.96, 0.96])

    col_widths = [0.06, 0.04, 0.04] + [0.06, 0.06] * 4 + [0.24, 0.06, 0.06]
    total_width = sum(col_widths)
    col_widths = [w / total_width for w in col_widths]

    n_cols = len(col_labels)
    table_bbox = [0, 0, 1, 0.82]
    table = ax.table(
        cellText=rows,
        colLabels=col_labels,
        cellLoc="center",
        colLoc="center",
        bbox=table_bbox,
    )
    table.auto_set_font_size(False)
    table.set_fontsize(9)

    for col_idx, width in enumerate(col_widths):
        for row_idx in range(n_rows + 1):
            cell = table.get_celld().get((row_idx, col_idx))
            if cell:
                cell.set_width(width)

    line_counts: List[int] = [1]
    for row in rows:
        max_lines = 1
        for value in row:
            max_lines = max(max_lines, str(value).count("\n") + 1)
        line_counts.append(max_lines)
    total_lines = sum(line_counts)
    bbox_height = table_bbox[3]
    for row_idx, line_count in enumerate(line_counts):
        row_height = bbox_height * (line_count / total_lines)
        for col_idx in range(n_cols):
            cell = table.get_celld().get((row_idx, col_idx))
            if cell:
                cell.set_height(row_height)
                if row_idx == 0:
                    cell.get_text().set_fontweight("bold")

    rows_problem = [row[0] for row in rows]
    for idx in range(1, n_rows):
        if rows_problem[idx] != rows_problem[idx - 1]:
            cell = table.get_celld().get((idx + 1, 0))
            if cell:
                ax.hlines(cell.get_y(), 0, 1, transform=ax.transAxes, color="#666666", linewidth=0.6)

    x_edges = [0.0]
    for width in col_widths:
        x_edges.append(x_edges[-1] + width)

    def center_between(start: int, end: int) -> float:
        return (x_edges[start] + x_edges[end]) / 2

    header_bottom = table_bbox[1] + table_bbox[3]
    header2_h = 0.06
    header1_h = 0.07
    header2_y = header_bottom
    header1_y = header_bottom + header2_h

    def draw_header_cell(x0: float, x1: float, y0: float, h: float, text: str, fontsize: int) -> None:
        rect = plt.Rectangle(
            (x0, y0),
            x1 - x0,
            h,
            fill=False,
            linewidth=0.8,
            edgecolor="#333333",
            transform=ax.transAxes,
        )
        ax.add_patch(rect)
        if text:
            ax.text(
                (x0 + x1) / 2,
                y0 + h / 2,
                text,
                ha="center",
                va="center",
                fontsize=fontsize,
                transform=ax.transAxes,
            )

    draw_header_cell(x_edges[0], x_edges[3], header1_y, header1_h, "Instances", 11)
    draw_header_cell(x_edges[3], x_edges[11], header1_y, header1_h, "Methods", 11)
    draw_header_cell(x_edges[11], x_edges[14], header1_y, header1_h, "Best method (others)", 11)

    draw_header_cell(x_edges[0], x_edges[3], header2_y, header2_h, "", 9)
    method_labels = ["SVGD", "PBIL", "MIMIC", "BOA"]
    method_spans = [(3, 5), (5, 7), (7, 9), (9, 11)]
    for label, (start, end) in zip(method_labels, method_spans):
        draw_header_cell(x_edges[start], x_edges[end], header2_y, header2_h, label, 9)
    draw_header_cell(x_edges[11], x_edges[14], header2_y, header2_h, "", 9)

    ax.hlines(header_bottom, 0, 1, transform=ax.transAxes, color="#333333", linewidth=0.8)

    output_png.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_png)
    fig.savefig(output_pdf)
    plt.close(fig)
    print(f"Saved table to {output_png} and {output_pdf}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate summary table for all instances.")
    parser.add_argument("--png", type=Path, default=ROOT / "courbes" / "summary_table.png")
    parser.add_argument("--pdf", type=Path, default=ROOT / "courbes" / "summary_table.pdf")
    parser.add_argument("--csv", type=Path, default=ROOT / "courbes" / "summary_table.csv")
    parser.add_argument("--xlsx", type=Path, default=ROOT / "courbes" / "summary_table.xlsx")
    parser.add_argument("--tex", type=Path, default=ROOT / "courbes" / "summary_table.tex")
    parser.add_argument("--format", choices=("all", "csv", "image", "excel", "tex"), default="all")
    args = parser.parse_args()

    # Ask for config
    config_name = input("Enter config name (e.g., krbf__advglobalrankweighted__M7__L13__eps0p08__g0p015__ds0p03__dm0p01): ").strip()
    config_dir = ROOT / "results" / "config" / config_name
    
    if not config_dir.exists():
        print(f"Error: Config directory not found: {config_dir}")
        return

    methods = ["PBIL", "MIMIC", "BOA"]
    rows, col_labels, raw_svgd_scores, best_other_map = build_rows(methods, config_dir)
    if args.format in ("all", "csv"):
        write_csv(rows, col_labels, args.csv)
    if args.format in ("all", "image"):
        plot_table(rows, col_labels, args.png, args.pdf)
    if args.format in ("all", "excel"):
        write_excel(rows, col_labels, args.xlsx)
    if args.format in ("all", "tex"):
        write_latex_table(
            rows,
            args.tex,
            raw_svgd_scores=raw_svgd_scores,
            config_dir=config_dir,
            best_other_map=best_other_map,
        )


if __name__ == "__main__":
    main()
